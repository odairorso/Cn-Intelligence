import zipfile
import sys

def main():
    print("Testing if fixed_fluxo.xlsx can be opened as a zip in Python...")
    try:
        with zipfile.ZipFile('fixed_fluxo.xlsx', 'r') as z:
            print("Zip file contents:")
            for name in z.namelist()[:10]:
                print("  ", name)
            print("Success opening as zip!")
    except Exception as e:
        print("Error opening zip:", e)

    # Let's try importing openpyxl
    try:
        import openpyxl
        print("openpyxl is installed. Opening workbook...")
        wb = openpyxl.load_workbook('fixed_fluxo.xlsx', read_only=True)
        print("Sheets:", wb.sheetnames)
        
        for name in wb.sheetnames:
            sheet = wb[name]
            # Search for BATISTOTE
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                row_str = str(row).upper()
                if 'BATISTOTE' in row_str:
                    print(f"[{name}] [Row {row_idx}] {row}")
    except Exception as e:
        print("Error with openpyxl:", e)

if __name__ == '__main__':
    main()
